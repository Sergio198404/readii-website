#!/usr/bin/env python3
"""
Readii Content Scanner
扫描内容文件夹，自动识别书籍和音频文件，生成 Excel 导入模板
Usage: python tools/scan_content.py --root "内容文件夹路径"
"""

import os
import re
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("请先安装依赖: pip install openpyxl")
    exit(1)

# ── 系列配置 ──
SERIES_CONFIG = {
    "GK": {"name": "Heinemann GK", "level": 1, "age_min": 3, "age_max": 5},
    "G1": {"name": "Heinemann G1", "level": 2, "age_min": 5, "age_max": 7},
    "G2": {"name": "Heinemann G2", "level": 3, "age_min": 6, "age_max": 8},
    "history": {"name": "History Series", "level": 4, "age_min": 7, "age_max": 12},
    "US":  {"name": "Science Series", "level": 4, "age_min": 7, "age_max": 12},
}

def classify_file(filename: str) -> str:
    """判断文件类型"""
    name = filename.lower()
    stem = Path(filename).stem.lower()

    if name.endswith('.pdf'):
        return 'pdf'
    if name.endswith('.docx') or name.endswith('.doc'):
        return 'docx'
    if not (name.endswith('.mp3') or name.endswith('.mp4') or name.endswith('.m4a')):
        return 'other'

    # V&G / G&V = 词汇讲解
    if 'v&g' in stem or 'g&v' in stem or 'vg' in stem:
        return 'commentary'
    # -Q = 问题音频
    if stem.endswith('-q') or '-q.' in stem:
        return 'question'
    # _1_1 / _1_2 = 多天阅读（_1_天数）
    if re.search(r'_1_\d+', stem):
        day = re.search(r'_1_(\d+)', stem)
        return f'reading_day_{day.group(1)}'
    # -1 / -2 结尾 = 阅读/互动
    if re.search(r'-1$', stem):
        return 'reading_day_1'
    if re.search(r'-2$', stem):
        return 'reading_day_2'
    if re.search(r'-3$', stem):
        return 'reading_day_3'
    if re.search(r'-4$', stem):
        return 'reading_day_4'
    # 其他 mp3 = 阅读
    return 'reading_day_1'

def clean_book_title(folder_name: str) -> str:
    """从文件夹名提取干净的书名"""
    # 去掉前缀如 D38-, 5-D38-050, D42-037 等
    name = re.sub(r'^[\d]+-D\d+-\d+\s*', '', folder_name)
    name = re.sub(r'^D\d+-\d+\s*', '', name)
    name = re.sub(r'^D\d+\s*', '', name)
    name = re.sub(r'^\d+-\w+-\w+\s*', '', name)
    return name.strip() or folder_name

def scan_series(root: Path, series_key: str) -> list:
    """扫描一个系列文件夹，返回书籍列表"""
    series_path = root / series_key
    if not series_path.exists():
        print(f"  跳过: {series_key} (文件夹不存在)")
        return []

    config = SERIES_CONFIG.get(series_key, {
        "name": series_key, "level": 1, "age_min": 3, "age_max": 12
    })
    books = []

    for book_folder in sorted(series_path.iterdir()):
        if not book_folder.is_dir():
            continue

        book_title = clean_book_title(book_folder.name)
        files = {
            'pdf': [], 'docx': [], 'commentary': [],
            'reading_day_1': [], 'reading_day_2': [],
            'reading_day_3': [], 'reading_day_4': [],
            'question': [], 'other': []
        }

        for f in sorted(book_folder.iterdir()):
            if f.name.startswith('.'):
                continue
            ftype = classify_file(f.name)
            if ftype in files:
                files[ftype].append(f.name)

        # 计算总天数
        total_days = 1
        for d in [2, 3, 4]:
            if files[f'reading_day_{d}']:
                total_days = d

        books.append({
            'series_key': series_key,
            'series_name': config['name'],
            'level': config['level'],
            'age_min': config['age_min'],
            'age_max': config['age_max'],
            'folder_name': book_folder.name,
            'book_title': book_title,
            'total_days': total_days,
            'pdf': files['pdf'][0] if files['pdf'] else '',
            'docx': files['docx'][0] if files['docx'] else '',
            'commentary': files['commentary'][0] if files['commentary'] else '',
            'reading_day_1': files['reading_day_1'][0] if files['reading_day_1'] else '',
            'reading_day_2': files['reading_day_2'][0] if files['reading_day_2'] else '',
            'reading_day_3': files['reading_day_3'][0] if files['reading_day_3'] else '',
            'reading_day_4': files['reading_day_4'][0] if files['reading_day_4'] else '',
            'question': files['question'][0] if files['question'] else '',
            'needs_review': '',
        })

    return books

def write_excel(all_books: list, output_path: Path):
    """生成 Excel 导入模板"""
    wb = openpyxl.Workbook()

    # ── Sheet 1: 所有书籍 ──
    ws = wb.active
    ws.title = "books_all"

    headers = [
        'series_key', 'series_name', 'level', 'age_min', 'age_max',
        'folder_name', 'book_title', 'total_days',
        'pdf', 'commentary', 'question', 'docx',
        'reading_day_1', 'reading_day_2', 'reading_day_3', 'reading_day_4',
        'needs_review', 'notes'
    ]

    header_fill = PatternFill("solid", fgColor="1B3A2D")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    warn_fill = PatternFill("solid", fgColor="FFF3CD")
    ok_fill = PatternFill("solid", fgColor="E8F2EC")

    for row, book in enumerate(all_books, 2):
        values = [book.get(h, '') for h in headers]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)

        # 标黄缺失文件的行
        missing = []
        if not book['pdf']: missing.append('PDF')
        if not book['reading_day_1']: missing.append('Reading')
        if not book['commentary']: missing.append('Commentary')

        if missing:
            ws.cell(row=row, column=headers.index('needs_review')+1).value = f"Missing: {', '.join(missing)}"
            for col in range(1, len(headers)+1):
                ws.cell(row=row, column=col).fill = warn_fill
        else:
            ws.cell(row=row, column=headers.index('needs_review')+1).fill = ok_fill

    # 调整列宽
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 30
    for col_letter in ['I','J','K','L','M','N','O','P']:
        ws.column_dimensions[col_letter].width = 35

    # ── Sheet 2: 问题汇总 ──
    ws2 = wb.create_sheet("needs_review")
    ws2.append(["folder_name", "book_title", "series", "missing_files"])
    ws2_fill = PatternFill("solid", fgColor="1B3A2D")
    for cell in ws2[1]:
        cell.fill = ws2_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for book in all_books:
        missing = []
        if not book['pdf']: missing.append('PDF')
        if not book['reading_day_1']: missing.append('Reading audio')
        if not book['commentary']: missing.append('Commentary audio')
        if missing:
            ws2.append([
                book['folder_name'], book['book_title'],
                book['series_name'], ', '.join(missing)
            ])

    # ── Sheet 3: 统计 ──
    ws3 = wb.create_sheet("summary")
    series_count = {}
    for book in all_books:
        s = book['series_name']
        series_count[s] = series_count.get(s, 0) + 1

    ws3.append(["Series", "Book Count"])
    for s, c in series_count.items():
        ws3.append([s, c])
    ws3.append(["TOTAL", len(all_books)])

    wb.save(output_path)
    print(f"\n✅ Excel 已生成: {output_path}")

def main():
    parser = argparse.ArgumentParser(description='Readii Content Scanner')
    parser.add_argument('--root', required=True, help='内容根目录路径')
    parser.add_argument('--output', default='tools/readii_content_import.xlsx', help='输出 Excel 路径')
    args = parser.parse_args()

    root = Path(args.root)
    if not root.exists():
        print(f"❌ 路径不存在: {root}")
        return

    print(f"📂 扫描根目录: {root}")
    all_books = []

    for series_key in SERIES_CONFIG.keys():
        books = scan_series(root, series_key)
        print(f"  {series_key}: {len(books)} 本书")
        all_books.extend(books)

    print(f"\n📚 共找到 {len(all_books)} 本书")

    output_path = Path(args.output)
    output_path.parent.mkdir(exist_ok=True)
    write_excel(all_books, output_path)

    # 打印缺失文件汇总
    missing_books = [b for b in all_books if not b['reading_day_1'] or not b['pdf']]
    if missing_books:
        print(f"\n⚠️  {len(missing_books)} 本书有缺失文件，见 Excel 的 needs_review sheet")

if __name__ == '__main__':
    main()
